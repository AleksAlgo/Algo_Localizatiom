"""
Jerome 1.0 — translation pipeline orchestrator.

Pipeline:
  Phase 1 — translate each file independently (Sonnet)
  Phase 2 — single cross-file Opus review (consistency-aware)
  Phase 3 — single QA evaluation across all files (Sonnet)
  Result  — per-file output files + single overall score
"""
from __future__ import annotations
import datetime
import json
import sys
from pathlib import Path

AGENT_SCRIPTS = Path(__file__).parent / "scripts"
sys.path.insert(0, str(AGENT_SCRIPTS))

from file_handlers import get_handler
import xliff_handler
from translator import translate_segments
from qa_judge import evaluate
from llm_clients import call_llm, parse_json_response
from report_generator import write_report

TRANSLATE_MODEL = {
    "provider": "openrouter",
    "model_id": "anthropic/claude-sonnet-4-5",
}
REVIEW_MODEL = {
    "provider": "openrouter",
    "model_id": "anthropic/claude-opus-4-5",
}
JUDGE_MODEL = {
    "provider": "openrouter",
    "model_id": "anthropic/claude-sonnet-4-5",
}

WORD_LIMIT = 6000

SEV_COLOR = {
    "Critical": "C00000",
    "Major":    "FF0000",
    "Minor":    "FFC000",
    "Neutral":  "808080",
}
DECISION_COLOR = {"PASS": "70AD47", "FAIL": "C00000"}


def write_excel_report(
    overall_qa: dict,
    results: list[dict],
    src_lang: str,
    tgt_lang: str,
    out_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── helpers ────────────────────────────────────────────────────────────
    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _bold(size=11, color="000000") -> Font:
        return Font(bold=True, size=size, color=color, name="Arial")

    def _normal(size=10) -> Font:
        return Font(size=size, name="Arial")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, col, text, bg="1F3864", fg="FFFFFF", size=11):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, size=size, color=fg, name="Arial")
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        return c

    def _cell(ws, row, col, value, bold=False, bg=None, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = _bold() if bold else _normal()
        if bg:
            c.fill = _fill(bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = border
        return c

    # ── Sheet 1: Сводка ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 40
    ws1.row_dimensions[1].height = 30

    date_str = datetime.date.today().isoformat()
    score = overall_qa.get("quality_score", 0)
    decision = overall_qa.get("decision", {})
    dec_label = decision.get("label", "") if isinstance(decision, dict) else str(decision)
    dec_key   = decision.get("decision", "FAIL") if isinstance(decision, dict) else "FAIL"
    word_count = overall_qa.get("word_count", 0)

    ws1.merge_cells("A1:B1")
    title_cell = ws1["A1"]
    title_cell.value = "Jerome 1.1 — Отчёт QA"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    title_cell.fill = _fill("1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    meta = [
        ("Дата",            date_str),
        ("Языковая пара",   f"{src_lang} → {tgt_lang}"),
        ("Слов всего",      word_count),
        ("Модель перевода", f"openrouter/{TRANSLATE_MODEL['model_id']}"),
        ("Модель ревью",    f"openrouter/{REVIEW_MODEL['model_id']}"),
        ("Модель QA",       f"openrouter/{JUDGE_MODEL['model_id']}"),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        _cell(ws1, i, 1, k, bold=True, bg="D9E1F2")
        _cell(ws1, i, 2, v)

    # Score block
    score_row = len(meta) + 3
    ws1.merge_cells(f"A{score_row}:B{score_row}")
    _hdr(ws1, score_row, 1, "Итоговый результат", bg="1F3864", size=12)
    ws1.row_dimensions[score_row].height = 22

    score_bg = "70AD47" if score < 20 else ("FFC000" if score < 30 else "C00000")
    _cell(ws1, score_row + 1, 1, "MQM Score", bold=True, bg="D9E1F2")
    _cell(ws1, score_row + 1, 2, round(score, 2), align="center", bg=score_bg)
    ws1.cell(row=score_row + 1, column=2).font = Font(bold=True, size=12, color="FFFFFF", name="Arial")

    _cell(ws1, score_row + 2, 1, "Решение", bold=True, bg="D9E1F2")
    dec_cell = _cell(ws1, score_row + 2, 2, dec_label, align="center", bg=DECISION_COLOR.get(dec_key, "808080"))
    dec_cell.font = Font(bold=True, color="FFFFFF", name="Arial")

    # Per-file table
    tbl_row = score_row + 4
    ws1.merge_cells(f"A{tbl_row}:B{tbl_row}")
    _hdr(ws1, tbl_row, 1, "Файлы", size=11)
    ws1.row_dimensions[tbl_row].height = 20

    hdrs = ["Файл", "Слов", "MQM Score", "Статус"]
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 14
    col_widths = [40, 10, 14, 20]
    for ci, (h, w) in enumerate(zip(hdrs, col_widths), start=1):
        _hdr(ws1, tbl_row + 1, ci, h, bg="2E75B6")
        ws1.column_dimensions[get_column_letter(ci)].width = w

    for fi, r in enumerate(results):
        row = tbl_row + 2 + fi
        qa = r.get("qa_result", {})
        fs = round(qa.get("quality_score", 0), 2)
        has_crit = any(e.get("severity", "").lower() == "critical" for e in qa.get("errors", []))
        fd = _make_decision(fs, has_crit)
        bg = "E2EFDA" if fd["decision"] == "PASS" else "FCE4D6"
        _cell(ws1, row, 1, r["file_name"], bg=bg)
        _cell(ws1, row, 2, r["n_words"], align="right", bg=bg)
        _cell(ws1, row, 3, fs, align="right", bg=bg)
        _cell(ws1, row, 4, fd["label"], align="center", bg=bg)

    # ── Sheet 2: Ошибки ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Ошибки")
    ws2.sheet_view.showGridLines = False

    err_hdrs = ["Файл", "Категория", "Тип", "Уровень", "Штраф", "Источник", "Перевод", "Комментарий"]
    err_widths = [24, 14, 18, 12, 8, 30, 30, 45]
    for ci, (h, w) in enumerate(zip(err_hdrs, err_widths), start=1):
        _hdr(ws2, 1, ci, h, bg="1F3864")
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 20

    err_row = 2
    for r in results:
        qa = r.get("qa_result", {})
        for err in qa.get("errors", []):
            sev = err.get("severity", "Minor")
            bg = SEV_COLOR.get(sev, "FFFFFF") + "33"  # light tint
            row_bg = None  # just use default for data rows
            _cell(ws2, err_row, 1, r["file_name"])
            _cell(ws2, err_row, 2, err.get("category", ""))
            _cell(ws2, err_row, 3, err.get("type", ""))
            sev_cell = _cell(ws2, err_row, 4, sev, align="center",
                             bg=SEV_COLOR.get(sev, "CCCCCC"))
            sev_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            _cell(ws2, err_row, 5, err.get("weight", 0), align="right")
            _cell(ws2, err_row, 6, err.get("source", ""))
            _cell(ws2, err_row, 7, err.get("target", ""))
            _cell(ws2, err_row, 8, err.get("comment", ""))
            ws2.row_dimensions[err_row].height = 40
            err_row += 1

    # ── Sheet 3: Разбивка ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Разбивка")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 12

    breakdown = overall_qa.get("score_breakdown", {})
    by_cat = breakdown.get("by_category", {})
    by_sev = breakdown.get("by_severity", {})

    _hdr(ws3, 1, 1, "По категориям", bg="1F3864")
    _hdr(ws3, 1, 2, "Штраф", bg="1F3864")
    for i, (k, v) in enumerate(sorted(by_cat.items(), key=lambda x: -x[1]), start=2):
        _cell(ws3, i, 1, k)
        _cell(ws3, i, 2, v, align="right")

    offset = len(by_cat) + 3
    _hdr(ws3, offset, 1, "По уровням", bg="2E75B6")
    _hdr(ws3, offset, 2, "Штраф", bg="2E75B6")
    for i, (k, v) in enumerate(by_sev.items(), start=offset + 1):
        sev_cell = _cell(ws3, i, 1, k)
        _cell(ws3, i, 2, v, align="right")
        sev_cell.fill = _fill(SEV_COLOR.get(k, "CCCCCC"))
        sev_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    wb.save(out_path)


def _make_decision(score: float, has_critical: bool) -> dict:
    if has_critical:
        return {"decision": "FAIL", "label": "Critical errors present"}
    if score < 10:
        return {"decision": "PASS", "label": "Excellent"}
    if score < 20:
        return {"decision": "PASS", "label": "Good"}
    if score < 30:
        return {"decision": "PASS", "label": "Acceptable"}
    if score < 50:
        return {"decision": "FAIL", "label": "Poor"}
    return {"decision": "FAIL", "label": "Unacceptable"}

SYSTEM_CROSS_REVIEW = """You are a senior translation editor reviewing draft translations of multiple related educational documents from {src_lang} to {tgt_lang}.

Your primary goal: ensure CONSISTENCY across all files — same terms, same register, same style — while fixing accuracy, fluency, and locale issues.

Rules:
1. Use consistent terminology across all segments regardless of which file they belong to.
2. Preserve all placeholders/tags exactly: {{0}}, %s, <b>, ${{var}}.
3. Do NOT translate brand names or code identifiers.
4. Each segment has a namespaced id like "file0_segId" — preserve these exactly.

For EACH segment return:
  - "id": the original namespaced segment id
  - "text": final reviewed translation
  - "changed": true/false
  - "reason": short reason (only if changed)

Return JSON: {{"reviews": [{{"id":"...", "text":"...", "changed":true, "reason":"..."}}]}}
"""


def _count_words(text: str) -> int:
    return len(text.split())


def _is_xliff(path: Path) -> bool:
    return path.suffix.lower() in (".xliff", ".xlf")


def extract_segments(path: Path) -> list[dict]:
    if _is_xliff(path):
        return xliff_handler.extract(path)
    raw = get_handler(path).extract()
    return [{"id": s.id, "text": s.text, "ctx": s.ctx} for s in raw]


def count_file_words(path: Path) -> int:
    segs = extract_segments(path)
    return sum(_count_words(s["text"]) for s in segs)


def _namespace(segs: list[dict], prefix: str) -> list[dict]:
    return [{"id": f"{prefix}_{s['id']}", "text": s["text"], "ctx": s.get("ctx", "")} for s in segs]


def _unnamespace(d: dict[str, str], prefix: str) -> dict[str, str]:
    p = prefix + "_"
    return {k[len(p):]: v for k, v in d.items() if k.startswith(p)}


def _cross_review(
    all_segs: list[dict],
    all_drafts: dict[str, str],
    src_lang: str,
    tgt_lang: str,
    batch_size: int = 20,
) -> dict[str, str]:
    """
    Single Opus review pass across all namespaced segments.
    Returns reviewed translations keyed by namespaced IDs.
    """
    system = SYSTEM_CROSS_REVIEW.format(src_lang=src_lang, tgt_lang=tgt_lang)
    final: dict[str, str] = {}

    for i in range(0, len(all_segs), batch_size):
        batch = all_segs[i:i + batch_size]
        payload = {"segments": [
            {"id": s["id"], "source": s["text"], "draft_translation": all_drafts.get(s["id"], "")}
            for s in batch
        ]}
        user = json.dumps(payload, ensure_ascii=False, indent=2)
        raw = call_llm(
            REVIEW_MODEL["provider"], REVIEW_MODEL["model_id"],
            system, user, max_tokens=8192, json_mode=True,
        )
        try:
            parsed = parse_json_response(raw)
        except Exception as e:
            print(f"[cross_review] JSON parse error batch {i}: {e}")
            for s in batch:
                final[s["id"]] = all_drafts.get(s["id"], "")
            continue

        for r in parsed.get("reviews", []):
            sid = r.get("id")
            if sid:
                final[sid] = r.get("text", all_drafts.get(sid, ""))

    for s in all_segs:
        final.setdefault(s["id"], all_drafts.get(s["id"], ""))

    return final


def run_pipeline(
    file_list: list[dict],
    src_folder_id: str,
    dst_folder_id: str,
    src_lang: str,
    tgt_lang: str,
    work_dir: Path,
    drive_download_fn,
    drive_upload_fn,
    progress_cb=None,
) -> dict:
    """
    Three-phase pipeline:
      Phase 1: translate each file (Sonnet)
      Phase 2: cross-file Opus review (single pass, consistency-focused)
      Phase 3: unified QA evaluation → single overall score

    Returns:
      {
        "files": [{"file_name", "n_words", "output_path", "qa_report_path"}, ...],
        "overall_qa": {...},   # single QA result across all files
        "overall_score": float,
      }
    """
    def _cb(stage: str):
        if progress_cb:
            progress_cb(stage)

    total = len(file_list)
    file_data = []  # per-file: {name, local_path, segs, draft, n_words}

    # ── Phase 1: download + translate ────────────────────────────────────────
    for i, f in enumerate(file_list):
        _cb(f"[{i+1}/{total}] {f['name']} — скачивание")
        local_path = drive_download_fn(f["id"], f["name"], f["mimeType"], work_dir)

        _cb(f"[{i+1}/{total}] {f['name']} — извлечение")
        segs = extract_segments(local_path)
        n_words = sum(_count_words(s["text"]) for s in segs)
        hl_segs = [s for s in segs if (isinstance(s.get("ctx"), dict) and s["ctx"].get("hyperlink")) or ".hl" in s["id"]]
        _cb(f"  → сегментов: {len(segs)}, гиперссылок: {len(hl_segs)}, слов: {n_words}")

        _cb(f"[{i+1}/{total}] {f['name']} — перевод")
        draft = translate_segments(
            segs,
            TRANSLATE_MODEL["provider"],
            TRANSLATE_MODEL["model_id"],
            src_lang=src_lang,
            tgt_lang=tgt_lang,
            domain="education",
            batch_size=20,
        )

        file_data.append({
            "drive_meta": f,
            "local_path": local_path,
            "segs": segs,
            "draft": draft,
            "n_words": n_words,
        })
        _cb(f"[{i+1}/{total}] {f['name']} — переведено")

    # ── Phase 2: cross-file Opus review ──────────────────────────────────────
    _cb(f"Ревью всех файлов (Opus)…")

    all_segs_ns: list[dict] = []
    all_drafts_ns: dict[str, str] = {}

    for i, fd in enumerate(file_data):
        prefix = f"file{i}"
        ns_segs = _namespace(fd["segs"], prefix)
        all_segs_ns.extend(ns_segs)
        for s in ns_segs:
            all_drafts_ns[s["id"]] = fd["draft"].get(s["id"].split("_", 1)[1], "")

    reviewed_ns = _cross_review(all_segs_ns, all_drafts_ns, src_lang, tgt_lang)
    _cb("Ревью завершено")

    # ── Save reviewed files ───────────────────────────────────────────────────
    results = []
    for i, fd in enumerate(file_data):
        prefix = f"file{i}"
        reviewed = _unnamespace(reviewed_ns, prefix)

        stem = fd["local_path"].stem
        suffix = fd["local_path"].suffix
        out_path = work_dir / f"{stem}_reviewed_{tgt_lang}{suffix}"
        if _is_xliff(fd["local_path"]):
            xliff_handler.write(fd["local_path"], reviewed, out_path)
        else:
            get_handler(fd["local_path"]).write(reviewed, out_path)

        _cb(f"Загрузка: {fd['drive_meta']['name']}")
        try:
            drive_upload_fn(dst_folder_id, out_path)
        except Exception as e:
            _cb(f"Ошибка загрузки {fd['drive_meta']['name']}: {e}")

        results.append({
            "file_name": fd["drive_meta"]["name"],
            "n_words": fd["n_words"],
            "output_path": out_path,
            "segs": fd["segs"],
            "reviewed": reviewed,
        })

    # ── Phase 3: per-file QA with original IDs → weighted overall score ─────
    _cb("QA-оценка (единая по всем файлам)…")

    total_weighted_score = 0.0
    total_words = 0
    all_errors = []
    per_file_qa = []

    for r in results:
        _cb(f"QA: {r['file_name']}")
        qa = evaluate(
            r["segs"],
            r["reviewed"],
            JUDGE_MODEL["provider"],
            JUDGE_MODEL["model_id"],
            src_lang=src_lang,
            tgt_lang=tgt_lang,
        )
        r["qa_result"] = qa
        per_file_qa.append(qa)
        w = r["n_words"] or 1
        total_weighted_score += qa.get("quality_score", 0) * w
        total_words += w
        all_errors.extend(qa.get("errors", []))

    overall_score = round(total_weighted_score / total_words, 2) if total_words else 0.0
    has_critical = any(
        e.get("severity", "").lower() == "critical" for e in all_errors
    )

    overall_decision = _make_decision(overall_score, has_critical)

    # copy metadata fields from first per-file QA result
    _meta = per_file_qa[0] if per_file_qa else {}
    overall_qa = {
        "quality_score": overall_score,
        "decision": overall_decision,
        "errors": all_errors,
        "per_file": per_file_qa,
        "judge_model": _meta.get("judge_model", JUDGE_MODEL["model_id"]),
        "score_breakdown": _meta.get("score_breakdown", {}),
        "total_score": sum(q.get("total_score", 0) for q in per_file_qa),
        "word_count": total_words,
        "summary_notes": "",
    }
    _cb("QA завершена")

    # Write combined QA report
    qa_report_path = work_dir / "jerome_qa_report.md"
    file_names = ", ".join(r["file_name"] for r in results)
    write_report(
        overall_qa, qa_report_path, stage="v1",
        source_file=file_names,
        output_file=", ".join(str(r["output_path"].name) for r in results),
        source_lang=src_lang, target_lang=tgt_lang, domain="education",
        translation_model=f"openrouter/{TRANSLATE_MODEL['model_id']}",
    )

    xlsx_report_path = work_dir / "jerome_qa_report.xlsx"
    try:
        write_excel_report(overall_qa, results, src_lang, tgt_lang, xlsx_report_path)
        _cb("Excel-отчёт создан")
    except Exception as e:
        _cb(f"Excel-отчёт: ошибка — {e}")

    for r in results:
        r.pop("segs", None)
        r.pop("reviewed", None)
        r["qa_report_path"] = qa_report_path

    return {
        "files": results,
        "overall_qa": overall_qa,
        "overall_score": overall_score,
        "overall_label": overall_decision.get("label", ""),
    }
